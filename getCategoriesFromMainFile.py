from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np

# input workbook path
input_workbook_path = "expenses_2024.xlsx"
# create a new output workbook
output_wb = Workbook()
# load the excel workbook
input_wb = load_workbook(input_workbook_path)
# get the active (default) worksheets
input_ws = input_wb.active
output_ws = output_wb.active

# category column
categoryCol = 'B'

# go over the category column and save the categories into a list
for categoryRow in range(6, 105):
    if (input_ws[categoryCol + str(categoryRow)].value is not None) and ('סה"כ' not in input_ws[categoryCol + str(categoryRow)].value) \
    and ('פירוט' not in input_ws[categoryCol + str(categoryRow)].value) and ("אפיק החיסכון" not in input_ws[categoryCol + str(categoryRow)].value):
        output_ws.append((input_ws[categoryCol + str(categoryRow)].value,))

# save the workbook
output_wb.save("categories.xlsx")

# TODO: using each father-category and expense name, connect a category from the sub-categories group
# for that sub-category, paste the equivalent expense amount in the proper cell in the output-expense-sheet 
