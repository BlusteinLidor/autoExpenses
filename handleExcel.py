from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openAI import sortExpensesAI
from data import *

# initiating the dict of the expenses name and the expense
expensesDict = {}

def getExpenses(workbook_path):
    # load the excel workbook
    wb = load_workbook(workbook_path)
    # get the active (default) worksheet
    ws = wb.active
    # expense name column
    expenseNameCol = 'B'
    # category column
    categoryCol = 'C'
    # expense cost column
    expenseCol = 'F'
    # first expense row
    categoryRow = 5
    # 
    global expensesDict
    totalCost = 0
    # loop through the rows and get the expense name and the category, separated by a dash
    while(ws[expenseNameCol + str(categoryRow)].value is not None):
        expenseNameCell = ws[expenseNameCol + str(categoryRow)]
        expenseCostCell = ws[expenseCol + str(categoryRow)]
        expenseCategoryCell = ws[categoryCol + str(categoryRow)]

        # if the expense name is already in the dict, add the cost of the expense to the last cost
        if expenseNameCell.value in expensesDict:
            expensesDict[expenseNameCell.value][0] += expenseCostCell.value
            totalCost += expenseCostCell.value
        # else, add the expense name, it's cost and it's category
        else:
            expensesDict.update({expenseNameCell.value : [expenseCostCell.value, expenseCategoryCell.value]})
            totalCost += expenseCostCell.value
        # go to the next row
        categoryRow += 1
    
    file = open("expensesDict.txt", "w")

    expensesSorted = sortExpensesAI(expensesDict)

    file.write(str(expensesDict) + "\n" + expensesSorted + "\n Total cost: " + str(totalCost))
    file.close()

    return expensesSorted

#@TODO add the expenses to the final excel file - go through each line in chat's response, for each line, check the name of the expense and it's cost, add the cost to a
# dict with the total cost for each category. then, add the category's cost to the final excel sheet

def fillCells(outputWorkbookPath, openAIOutput, month):
    # load the excel workbook
    wb = load_workbook(outputWorkbookPath)
    # get the active (default) worksheet
    ws = wb.active
    # category column
    categoryCol = 'B'
    # expense column
    if month not in months:
        expenseCol = monthToExpenseColDict[12]
    else:
        expenseCol = monthToExpenseColDict[int(month)]

    # starting category row
    categoryRow = 17
    # max category row
    maxCategoryRow = 105
   # initiate a list of the expenses names and their associated category
    expenseNameCostCategoryDict = {}
    # split openai's output to separate lines
    eachLineList = openAIOutput.split("\n")
    # for each line
    for val in eachLineList:
        # for each expense - cost - category part, split between the expense, cost and category
        expenseNameCostCategory = val.rsplit(" - ", 2)
        # add each [expenseName, cost, category] tuple to a dict
        expenseNameCostCategoryDict.update({expenseNameCostCategory[0] : [expenseNameCostCategory[1], expenseNameCostCategory[2]]})

    for val in expenseNameCostCategoryDict:
        categoryRow = 17
        while(categoryRow < maxCategoryRow):
            # if the expense name in the expenseNameCostCategoryDict equals the expense name in the excel file
            if ws[categoryCol + str(categoryRow)].value is not None and ws[categoryCol + str(categoryRow)].value in expenseNameCostCategoryDict.get(val)[1]:
                #@TODO add the expense cost to the output excel file
                if ws[expenseCol + str(categoryRow)].value is None:
                    ws[expenseCol + str(categoryRow)].value = float(expenseNameCostCategoryDict.get(val)[0])
                else:
                    ws[expenseCol + str(categoryRow)].value = float(ws[expenseCol + str(categoryRow)].value) + float(expenseNameCostCategoryDict.get(val)[0])
                #@TODO add the name of the expense to the comment (append the name, don't replace it)
                if ws[expenseCol + str(categoryRow)].comment is None:
                    ws[expenseCol + str(categoryRow)].comment = Comment("", "Automated")
                #comment = Comment((ws[expenseCol + str(categoryRow)] ,"\n", val), "Automated")
                comment = Comment(val + " - " + str(expenseNameCostCategoryDict.get(val)[0]), "Automated")
                if comment.text != "":
                    ws[expenseCol + str(categoryRow)].comment.text += comment.text + "\n"
            categoryRow += 1

    wb.save(outputWorkbookPath[:-5] + "_output" + outputWorkbookPath[-5:])
                
# getExpenses(r"C:\Users\liido\Downloads\transaction-details_export_1728243409088.xlsx")

#string = """TRES JOLIE - 20 - רהיטים, כלי בית וגן  
#מאפיתת אורן משי באר שבע - 44 - מצרכי מזון  
#סופר פארם מצדה ב"ש - 46.8 - תרופות, בדיקות וטיפולים רפואיים  
#7 בעיר הבלוק - 39.8 - מצרכי מזון  
#סהרה אלקטרוניקה בע"מ - 184 - מתנות  
#חברת החשמל לישראל בע"מ - 368.73 - חשבון חשמל  
#פז אפליקציית יילו - 683.15 - חשבון גז  
#פרפל יבוא ושיווק - 77 - תכשיטים וקוסמטיקה  
#סופרמרקט הגשר - 15 - מצרכי מזון  
#אייבורי מחשבים - 195 - גאדג'טים ואלקטרוניקה  
#אפריל פארק הקרח - 127 - מוצרי היגיינה  
#חברת פרטנר תקשורת בע"מ (ה - 119 - סלולארי  
#מחסני השוק אינטרנט - 1350.04 - מצרכי מזון  
#היי קוקי - 122.55 - אוכל בחוץ (כולל מסעדות, בתי קפה, משלוחים)  
#כביש 6 - 76.38 - כביש 6 וחוצה צפון  
#WOLT - 320 - אוכל בחוץ (כולל מסעדות, בתי קפה, משלוחים)  
#סופר קופיקס - 63.4 - מצרכי מזון  
#פלאפון חשבון תקופתי - 41.84 - סלולארי  
#מ. התחבורה ר.רכב - 906 - דו"חות וקנסות  
#שוהם סטוק בע"מ - 58.74 - מצרכי מזון  
#AIG ביטוח רכב - 2855 - ביטוח רכב  
#המילטון חשמל ואלקטרוניקה - 450 - גאדג'טים ואלקטרוניקה  
#סיטי שופ - 79.28 - ספרים ומוזיקה  
#הסטוק גרנד קניון ב"ש~צמרת - 15.9 - מתנות  
#תכשיטי ישראל - 30 - מתנות  
#אושר עד באר שבע - 140.7 - מצרכי מזון  
#פנגו חשבונית חודשית - 9.32 - הוצאות תחבורה ציבורית ואופניים  
#מי שבע בע"מ - 114.02 - דו"חות וקנסות  
#רוזה & פולה - 40 - אוכל בחוץ (כולל מסעדות, בתי קפה, משלוחים)  
#דמי כרטיס - 0.0 - משיכת כספים (כספומט)  
#עיריית באר שבע - 390.58 - דו"חות וקנסות  
#BIT - 100 - העברת כספים  
#יעקב כהן - 60 - תספורת  
#מס הכנסה עצמאים וחברות - 500 - דו"חות וקנסות"""
#
#fillCells("C:/Users/liido/Downloads/expenses_2024.xlsx", string, 12)